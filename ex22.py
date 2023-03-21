import openpyxl

wb = openpyxl.load_workbook('input.xlsx')

ws = wb['MAU']

def gpaScore(math, physics, chemistry):
    return (math + physics + chemistry) / 3;

def sortByNameAToZ():

    sorted_rows = sorted(ws.rows, key=lambda x: x[2].value)

    wb_new = openpyxl.Workbook()
    ws_new = wb_new.active
    for row in sorted_rows:
        ws_new.append([cell.value for cell in row])

    wb_new.save('output.xlsx')

def statistics():
    countExcellent = 0 
    countGood= 0 
    countAverage= 0
    
    for row in ws.iter_rows(min_row=2):
        average_score = gpaScore(row[4], row[5], row[6]) 
        if average_score >= 8.0:
            countExcellent += 1 
        elif average_score >= 6.5:
            countGood += 1 
        else:
            average_count += 1 
    #vị trí dữ liệu thống kê
    ws['Q65'] = countExcellent
    ws['R65'] = countGood
    ws['S65'] = countAverage